import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import os
from config import EXCEL_EXPORT_PATH, IMAGE_DIR,CSV_PATH


#Por si las carpetas no  existen.
IMAGE_DIR.mkdir(parents=True, exist_ok=True)
EXCEL_EXPORT_PATH.parent.mkdir(parents=True, exist_ok=True)

# Cargar el CSV
df = pd.read_csv(CSV_PATH, 
        encoding='latin1',parse_dates=["Order Date", "Ship Date"])

#Eliminar  columnas innecesarias para el analisis
df = df.drop(columns=["Row ID", "Postal Code","Country","Region"])

#Verificar  que ya no estan
print(df.columns)

# Crear columnas adicionales para año y mes para facilitar analisis
df['Year'] = df['Order Date'].dt.year
df['Month'] = df['Order Date'].dt.month
df['Month_Name'] = df['Order Date'].dt.strftime('%B')  # Nombre del mes

#Aunque el dataset no tiene una  categoría para medir el tiempo de entrega, se puede  calcular así
df["Shipping_Days"] = (df["Ship Date"] - df["Order Date"]).dt.days


#Verificar  que  estan
print(df.columns)

#Agrupar por año  y  mes todas las  ganancias
monthly_Profit = df.groupby(['Year','Month','Month_Name'])["Profit"].sum().reset_index()

# Ordenar correctamente por año y mes
monthly_Profit = monthly_Profit.sort_values(by=['Year', 'Month'])

#Se elimina el numero del mes porque no es necesario
monthly_Profit=monthly_Profit.drop(columns=["Month"])

#Ventas y ganancia generada  por año
year_sales_and_profit=df.groupby("Year")[["Sales","Profit"]].sum().reset_index()

#Se crea columna  para conocer el porcentaje de ganancia por año
year_sales_and_profit["Profit Margin (%)"]=(year_sales_and_profit["Profit"] * 100) / year_sales_and_profit["Sales"]

#Agrupar las ventas y la ganancia por categoria
sales_and_profit_by_category = df.groupby("Category")[["Sales","Profit"]].sum().reset_index()

#Se crea columna  para conocer el porcentaje de ganancia por categoria
sales_and_profit_by_category["Profit Margin (%)"]=(sales_and_profit_by_category["Profit"] * 100) / sales_and_profit_by_category["Sales"]

#Agrupar las  ventas por sub-categoria
sales_and_profit_by_sub_category=df.groupby("Sub-Category")[["Sales","Profit"]].sum().reset_index()

#Se crea columna  para conocer el porcentaje de ganancia por Sub categoria
sales_and_profit_by_sub_category["Profit Margin (%)"]=(sales_and_profit_by_sub_category["Profit"] * 100) / sales_and_profit_by_sub_category["Sales"]


#Agrupar por cliente la cantidad total de compras (número de órdenes)
orders_by_customer=df.groupby(["Customer ID","Customer Name"])["Sales"].sum().reset_index()

#Agrupar clientes por ganancia
customers_by_profit=df.groupby(["Customer ID","Customer Name"])["Profit"].sum().reset_index()

#Unir las dos metricas para el analisis
customer_analysis = pd.merge(orders_by_customer, customers_by_profit, on=["Customer ID", "Customer Name"])

#Productos más vendidos (por cantidad)
top_productos=df.groupby("Product Name")["Quantity"].sum().reset_index()

#Solo mostrar los 10 productos mas vendidos
top_productos=top_productos.sort_values(by="Quantity",ascending=False).head(10)

# Todos los productos con sus ganancias sumadas
all_products_profit = df.groupby("Product Name")["Profit"].sum().reset_index()

# Los 10 productos con más ganancias
top_profitable_products = all_products_profit.sort_values(by="Profit", ascending=False).head(10)

#Solo mostrar los 10 productos con mas ganancias
top_profitable_products = top_profitable_products.sort_values(by="Profit", ascending=False).head(10)

# Los 10 productos con menos ganancias
least_profitable_products = all_products_profit.sort_values(by="Profit", ascending=True).head(10)


#Mostrar las 6 ciudades con mas ganancias
profit_by_city = df.groupby("City")["Profit"].sum().reset_index().sort_values(by="Profit", ascending=False).head(6)

#impacto del descuento en la ganancia promedio
discount_impact = df.groupby("Discount")["Profit"].mean().reset_index().sort_values(by="Discount")

#Tiempo  de entrega por metodo de envío y su ganancia
shipping_analysis = df.groupby("Ship Mode").agg({"Shipping_Days": "mean","Profit": "sum"}).reset_index()

#Agrupacion de clientes con la  suma total de compras y de las ganancias generadas.
customers_stats=df.groupby(["Customer ID","Customer Name"]).agg(
    {"Sales": "sum","Profit": "sum"}).reset_index()

#Clientes que gastan una suma considerable pero no generan  ganancias
unprofitable_customers=customers_stats[(customers_stats["Sales"]>8000) & (customers_stats["Profit"]<0)]

#Clientes que gastan una suma considerable pero no generan  ganancias
unprofitable_customers=customers_stats[(customers_stats["Sales"]>8000) & (customers_stats["Profit"]<0)]

#Cliente  que  generó  la mayor  pérdida  en ganancia
most_unprofitable_customer = df[(df["Customer ID"] == "GT-14635") &(df["Profit"]<0)][["Customer ID", "Sales", "Profit", "Product ID","Discount"]]

# ======================
# Gráficas
# ======================
def plot_monthly_profits(dataframe):
    """
    Función  para graficar las ganancias mensuales (profit)
    
    """
    
   
    plt.figure(figsize=(15, 6))
    plt.plot(dataframe['Profit'], marker='o', linewidth=2, color='green', markersize=8)
    
    plt.title('Monthly Profit Evolution', fontsize=16)
    plt.xlabel('Year', fontsize=12)
    plt.ylabel('Profit ($)', fontsize=12)
    
    # Encontrar donde cambia el año para poner las etiquetas
    year_positions = []
    years = []
    previous_year = None
    
    for i, current_year in enumerate(dataframe['Year']):
        if current_year != previous_year:  # Si cambió el año
            year_positions.append(i)       # Guardar la posición
            years.append(current_year)     # Guardar el año
            previous_year = current_year
    
    # Poner solo los años en el eje X
    plt.xticks(year_positions, years)
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(IMAGE_DIR/"monthly_profit_chart.png", bbox_inches="tight")
    plt.show()

# llamar a la función
plot_monthly_profits(monthly_Profit)


def plot_sales_and_profit_with_margin(dataframe): #This.
    """
    Función para graficar las ventas y las ganancias anuales junto con su porcentaje de ganancia(Profit)
    """
    x = np.arange(len(dataframe['Year']))
    width = 0.35

    plt.figure(figsize=(10, 6))

    # Barras
    bars_sales = plt.bar(x - width/2, dataframe['Sales'], width=width, label='Sales')
    bars_profit = plt.bar(x + width/2, dataframe['Profit'], width=width, label='Profit')

    # Anotar el porcentaje de ganancia sobre cada barra de Profit
    for i, bar in enumerate(bars_profit):
        height = bar.get_height()
        margin = dataframe["Profit Margin (%)"].iloc[i]
        plt.text(bar.get_x() + bar.get_width()/2, height + 5000, f'{margin:.1f}%', 
                 ha='center', va='bottom', fontsize=9, color='green')

    # Configuración general
    plt.xticks(x, dataframe['Year'])
    plt.title('Annual Sales and Profit with Profit Margin (%)')
    plt.xlabel('Year')
    plt.ylabel('Amount ($)')
    plt.legend()
    plt.tight_layout()
    plt.savefig(IMAGE_DIR/"sales_and_profit_with_margin.png", bbox_inches="tight")
    plt.show()

#llamar a la función
plot_sales_and_profit_with_margin(year_sales_and_profit)


def plot_sales_profit_by_category(dataframe): #This.
    """
    Función para graficar las ventas y las ganancias por categoria junto con su porcentaje de ganancia(Profit)
    """
    x = np.arange(len(dataframe['Category']))
    width = 0.35

    plt.figure(figsize=(10, 6))

    # Barras
    bars_sales = plt.bar(x - width/2, dataframe['Sales'], width=width, label='Sales',color='skyblue')
    bars_profit = plt.bar(x + width/2, dataframe['Profit'], width=width, label='Profit', color='mediumseagreen')

    # Anotar el porcentaje de ganancia sobre cada barra de Profit
    for i, bar in enumerate(bars_profit):
        height = bar.get_height()
        margin = dataframe["Profit Margin (%)"].iloc[i]
        plt.text(bar.get_x() + bar.get_width()/2, height + 5000, f'{margin:.1f}%', 
                 ha='center', va='bottom', fontsize=12, color='black')
    
    
    plt.xticks(x, dataframe['Category'])
    plt.title('Category Sales and Profit with Profit Margin (%)')
    plt.xlabel('Category')
    plt.ylabel('Amount ($)')
    plt.legend()
    plt.tight_layout()
    plt.savefig(IMAGE_DIR/"sales_profit_by_category.png", bbox_inches="tight")

    plt.show()

#llamar a la función
plot_sales_profit_by_category(sales_and_profit_by_category)



def plot_sales_and_profit_by_sub_category(dataframe): #This.
    """
    Función para graficar las ventas y las ganancias por subcategoria junto con su porcentaje de ganancia(Profit)
    """
    x = np.arange(len(dataframe['Sub-Category']))
    width = 0.35

    plt.figure(figsize=(14, 7))

    # Barras
    bars_sales = plt.bar(x - width/2, dataframe['Sales'], width=width, label='Sales',color='skyblue')
    bars_profit = plt.bar(x + width/2, dataframe['Profit'], width=width, label='Profit', color='mediumseagreen')

    # Anotar el porcentaje de ganancia sobre cada barra de Profit
    for i, bar in enumerate(bars_profit):
        height = bar.get_height()
        margin = dataframe["Profit Margin (%)"].iloc[i]
        plt.text(bar.get_x() + bar.get_width()/2+ 0.1, height + 5000, f'{margin:.1f}%', 
                 ha='center', va='bottom', fontsize=10, color='black')
    
    # Configuración general
    plt.xticks(x, dataframe['Sub-Category'],rotation=45, ha='right')
    plt.title('Sub-Category Sales and Profit with Profit Margin (%)')
    plt.xlabel('Sub-Category')
    plt.ylabel('Amount ($)')
    plt.legend()
    plt.tight_layout()
    plt.savefig(IMAGE_DIR/"sales_and_profit_by_sub_category.png", bbox_inches="tight")
    plt.show()

#llamar a la función
plot_sales_and_profit_by_sub_category(sales_and_profit_by_sub_category)



def plot_top_customers_by_profit(dataframe, top_n=10): 
    """
    Función para graficar el top 10 de Clientes que más han generado ganancias junto con la totalidad
    de dinero gastado
    """
    top_customers = dataframe.sort_values(by="Profit", ascending=False).head(top_n)

    plt.figure(figsize=(18, 6))
    bars = plt.barh(top_customers["Customer Name"], top_customers["Profit"], color='mediumseagreen')
    plt.xlabel("Profit ($)")
    plt.title(f"Top {top_n} Customers by Profit")

    # Ajustar límite del eje X para que el texto no se corte
    max_profit = top_customers["Profit"].max()
    plt.xlim(0, max_profit * 1.4)  

    for i, bar in enumerate(bars):
        sales = top_customers["Sales"].iloc[i] #Poner cuanto dinero se gastó el cliente
        customer_id = top_customers["Customer ID"].iloc[i] #Poner el ID para identificar al cliente(Por si hay nombres repetidos)

        plt.text(bar.get_width() + max_profit * 0.001,  
                 bar.get_y() + bar.get_height()/2,
                 f"{sales:.2f}$ Money spent| ID: {customer_id}", va='center', fontsize=10, color='black')
    plt.savefig(IMAGE_DIR/"top_customers_by_profit.png", bbox_inches="tight")
    plt.show()

#llamar a la función
plot_top_customers_by_profit(customer_analysis)


def plot_top_selling_products(dataframe):
    """
    Funcion para graficar los 10 productos mas vendidos.
    """
   
    dataframe = dataframe.sort_values(by="Quantity", ascending=True) #Para que el mas vendido quede de primero

    plt.figure(figsize=(14, 7))
    bars = plt.barh(dataframe["Product Name"], dataframe["Quantity"], color="steelblue")

    plt.xlabel("Quantity Sold")
    plt.title("Top 10 Most Sold Products")
    

    for i, bar in enumerate(bars):
        quantity = dataframe["Quantity"].iloc[i]
        plt.text(bar.get_width() + 1, bar.get_y() + bar.get_height() / 2,
                 f"{quantity}", va="center", fontsize=9)

    plt.tight_layout()
    plt.savefig(IMAGE_DIR/"top_selling_products.png", bbox_inches="tight")

    plt.show()

#Llamar a la función
plot_top_selling_products(top_productos)


def plot_top_profitable_products(dataframe): #This
    """
    Función que  grafica los  10 productos con más ganancia (profit)
    """

    # Ordenar para que el producto más rentable esté arriba
    dataframe = dataframe.sort_values(by="Profit", ascending=True)

    plt.figure(figsize=(14, 7))  
    bars = plt.barh(dataframe["Product Name"], dataframe["Profit"], color="mediumseagreen")

    plt.xlabel("Profit ($)")
    plt.title("Top 10 Most Profitable Products")

    
    max_profit = dataframe["Profit"].max()
    plt.xlim(0, max_profit * 1.25)  

    # Mostrar texto con las ganancias
    for i, bar in enumerate(bars):
        profit = dataframe["Profit"].iloc[i]
        plt.text(bar.get_width() + max_profit * 0.01,  # Texto un poco a la derecha
                 bar.get_y() + bar.get_height() / 2,
                 f"${profit:,.2f}", va="center", fontsize=9)

    plt.tight_layout()
    plt.savefig(IMAGE_DIR/"top_profitable_products.png", bbox_inches="tight")
    plt.show()

#llamar a la función
plot_top_profitable_products(top_profitable_products)



#Ejemplo útil para aprender a manejar gráficas con valores negativos.
def plot_least_profitable_products(dataframe): #This
    """
    Función para graficar los 10 productos menos rentables (con menos profit)
    """

    plt.figure(figsize=(14, 7))
    bars = plt.barh(dataframe["Product Name"], dataframe["Profit"], color="red")

    plt.xlabel("Profit ($)")
    plt.title("10 Least Profitable Products")

    min_profit = dataframe["Profit"].min()
    max_profit = dataframe["Profit"].max()

    plt.xlim(min_profit * 1.3, 0)  

    for i, bar in enumerate(bars):
        profit = dataframe["Profit"].iloc[i]
        
        xpos = bar.get_width() - abs(min_profit) * 0.05
        plt.text(xpos,
                 bar.get_y() + bar.get_height() / 2,
                 f"${profit:,.2f}",
                 va="center",
                 ha='right',  
                 fontsize=9,
                 color='black')

    plt.tight_layout()
    plt.savefig(IMAGE_DIR/"least_profitable_products.png", bbox_inches="tight")
    plt.show()

#llamar a la función
plot_least_profitable_products(least_profitable_products)



def plot_profit_by_city(dataframe): #This.
    """
    Función que  realiza una gráfica  pastel para mostrar
    las 6 ciudades que más generan ganancias.
    """
    # Seleccionar las 6 ciudades con más ganancias
    top_6_cities = dataframe.sort_values(by="Profit", ascending=False).head(6)
    
    plt.figure(figsize=(12, 8))
    plt.pie(top_6_cities["Profit"], 
            labels=top_6_cities["City"], 
            autopct='%1.1f%%', 
            startangle=140
            )
    plt.title("Top 6 Cities by Profit",pad=20,fontsize=16)
    plt.axis('equal')  
    plt.savefig(IMAGE_DIR/"profit_by_city.png", bbox_inches="tight")
    plt.show()

#llamar a la función
plot_profit_by_city(profit_by_city)



def plot_discount_impact(dataframe): #This.
    """
    Función que grafica el impacto del descuento en la ganancia
    
    """
    # Crear una columna con descuento en porcentaje
    dataframe['Discount_&'] = dataframe['Discount'] * 100

    plt.figure(figsize=(10,6))
    plt.plot(dataframe['Discount_&'], dataframe['Profit'], marker='o', linestyle='-')
    
    plt.xlabel('Discount (%)')
    plt.ylabel('Average Profit ($)')
    plt.title('Impact of Discount on Average Profit')
    plt.grid(True)
    plt.savefig(IMAGE_DIR/"discount_impact.png", bbox_inches="tight")

    plt.show()

#llamar a la función
plot_discount_impact(discount_impact)



def plot_shipping_days_and_profit(dataframe): #This.
    """
    Función que crea un gráfico de barras para mostrar el tiempo promedio de entrega (en días) 
    por método de envío, junto con la ganancia total asociada a cada uno.
    """
    x = np.arange(len(dataframe['Ship Mode']))
    width = 0.3

    plt.figure(figsize=(10, 6))

    # Barra de Shipping Days
    bars = plt.bar(x, dataframe['Shipping_Days'], width=width, color='green', label='Avg Shipping Days')

    # Anotar Profit encima de cada barra
    for i, bar in enumerate(bars):
        shipping_days = bar.get_height()
        profit = dataframe["Profit"].iloc[i]
        plt.text(bar.get_x() + bar.get_width()/2, shipping_days + 0.1,
                 f'${profit:,.0f}', ha='center', va='bottom', fontsize=10.5, color='black')

    plt.xticks(x, dataframe['Ship Mode'])
    plt.ylabel('Average Shipping Days')
    plt.title('Average Shipping Time and Profit by Ship Mode')
    plt.tight_layout()
    plt.savefig(IMAGE_DIR/"shipping_days_and_profit.png", bbox_inches="tight")

    plt.show()

#llamar a la función
plot_shipping_days_and_profit(shipping_analysis)



def plot_unprofitable_customers(dataframe):#This.
    """
    Función que  grafica a clientes con alto dinero gastado
    pero con poca ganancia (profit)
    """
    plt.figure(figsize=(12, 6))
    bars = plt.barh(dataframe["Customer Name"], dataframe["Profit"], color='red')
    plt.xlabel("Profit ($)")
    plt.title("High-Spending Customers with Negative Profit")

    min_profit = dataframe["Profit"].min()
    plt.xlim(min_profit * 1.4, 0)  

    for i, bar in enumerate(bars):
        profit = dataframe["Profit"].iloc[i]
        sales = dataframe["Sales"].iloc[i]
        customer_id = dataframe["Customer ID"].iloc[i]
        plt.text(bar.get_width() - abs(min_profit) * 0.05, 
                 bar.get_y() + bar.get_height() / 2,
                 f"ID: {customer_id} | Spent: ${sales:,.0f}",
                 va='center', ha='right', fontsize=9, color='black')

    plt.tight_layout()
    plt.savefig(IMAGE_DIR/"unprofitable_customers.png", bbox_inches="tight")

    plt.show()

#llamar a la función
plot_unprofitable_customers(unprofitable_customers)



def plot_most_unprofitable_customer(df): #Util para manejar el Zip.

    """
    Función que grafica lo que  ha comprado el cliente que mas perdidas le  ha generado  a la empresa.

    """
    product_ID = df["Product ID"]
    profits = df["Profit"]
    discounts = df["Discount"]
    
    
    plt.figure(figsize=(12, 6))
    bars = plt.barh(product_ID, profits, color="red",height=0.3)
    
    plt.title("Negative Profit by Product - Customer GT-14635", fontsize=14)
    plt.xlabel("Profit ($)")
    plt.ylabel("Product ID")
    
    
    min_profit = profits.min()
    plt.xlim(min_profit * 1.4, 0) 
    
    
    for bar, discount in zip(bars, discounts):
        plt.text(bar.get_width() - abs(min_profit) * 0.05,  
                 bar.get_y() + bar.get_height() / 2,
                 f"{discount:.0%} discount",
                 va='center',
                 ha='right',
                 fontsize=12,
                 color='black')
    
    plt.tight_layout() 
    plt.savefig(IMAGE_DIR/"most_unprofitable_customer.png", bbox_inches="tight")
    plt.show()

#llamar a la función
plot_most_unprofitable_customer(most_unprofitable_customer)


# ======================
# Exportar datos a Excel
# ======================



import os
from openpyxl import Workbook

def add_sheet_with_data_and_image(df, excel_path, sheet_name, image_path, image_cell="G2"):
    """
    Agrega una hoja con datos y una imagen a un archivo Excel. Si el archivo no existe, lo crea.
    """

    # Crear el archivo si no existe
    if not os.path.exists(excel_path):
        wb = Workbook()
        wb.save(excel_path)

    
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)


   
    book = load_workbook(excel_path)
    ws = book[sheet_name]
    img = XLImage(image_path)
    ws.add_image(img, image_cell)
    book.save(excel_path)




#Llamada a la  función para  exportar  todos los  datos.
add_sheet_with_data_and_image(monthly_Profit,
    EXCEL_EXPORT_PATH,
    "monthly_profit",
    IMAGE_DIR / "monthly_profit_chart.png")

add_sheet_with_data_and_image(year_sales_and_profit,
    EXCEL_EXPORT_PATH,
    "year_sales_profit",
    IMAGE_DIR / "sales_and_profit_with_margin.png")

add_sheet_with_data_and_image(sales_and_profit_by_category,
    EXCEL_EXPORT_PATH,
    "sales_profit_category",
    IMAGE_DIR / "sales_profit_by_category.png")

add_sheet_with_data_and_image(sales_and_profit_by_sub_category,
    EXCEL_EXPORT_PATH,
    "sales_profit_subcategory",
    IMAGE_DIR / "sales_and_profit_by_sub_category.png")

add_sheet_with_data_and_image(customer_analysis,
    EXCEL_EXPORT_PATH,
    "top_customers_profit",
    IMAGE_DIR / "top_customers_by_profit.png")

add_sheet_with_data_and_image(top_productos,
    EXCEL_EXPORT_PATH,
    "top_products",
    IMAGE_DIR / "top_selling_products.png")

add_sheet_with_data_and_image(top_profitable_products,
    EXCEL_EXPORT_PATH,
    "top_profitable_prod",
    IMAGE_DIR / "top_profitable_products.png")

add_sheet_with_data_and_image(least_profitable_products,
    EXCEL_EXPORT_PATH,
    "least_profit_prod",
    IMAGE_DIR / "least_profitable_products.png")

add_sheet_with_data_and_image(profit_by_city,
    EXCEL_EXPORT_PATH,
    "profit_by_city",
    IMAGE_DIR / "profit_by_city.png")

add_sheet_with_data_and_image(discount_impact,
    EXCEL_EXPORT_PATH,
    "discount_impact",
    IMAGE_DIR / "discount_impact.png")

add_sheet_with_data_and_image(shipping_analysis,
    EXCEL_EXPORT_PATH,
    "ship_days_profit",
    IMAGE_DIR / "shipping_days_and_profit.png")

add_sheet_with_data_and_image(unprofitable_customers,
    EXCEL_EXPORT_PATH,
    "unprofitable_cust",
    IMAGE_DIR / "unprofitable_customers.png")

add_sheet_with_data_and_image(most_unprofitable_customer,
    EXCEL_EXPORT_PATH,
    "most_unprofit_cust",
    IMAGE_DIR / "most_unprofitable_customer.png")

